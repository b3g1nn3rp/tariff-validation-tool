#############################################################################################
###   Author - Priyesh Bhardwaj							          ###
###   Created on - 24/04/2020								  ###
###   Purpose - Validation of tariff charges populated on ape1_rated_events	          ###
###             as per rate card 							  ###
###   Language - Python3						                  ###
#############################################################################################


import openpyxl as xl
import math
import datetime
import pdb
import time
import logging
from colorama import Fore
from tqdm import *
start_time = time.time()

logging.basicConfig(filename=(f'C:\\Users\\ASHISBHA\\Desktop\\Tariff Validation Tool\\Logs\\tariff_validation_tool_{start_time}.log'), filemode='a', format='%(asctime)s- %(name)s - %(levelname)s - %(message)s',datefmt='%d-%b-%y %H:%M:%S',level=logging.DEBUG)



def load_Excel_1(filename,sheet_name):

	workbook = xl.load_workbook(filename)
	no_of_sheets = len(workbook.sheetnames)
	for i in range(0,no_of_sheets):
		if  workbook.sheetnames[i] == sheet_name:
			logging.info('Excel %s loaded successfully',filename)
			print((f'Sheet {sheet_name} of Excel {filename} loaded successfully'))
			return workbook,workbook.worksheets[i]

def load_Excel(filename,sheet_name1,sheet_name2):
	
	workbook = xl.load_workbook(filename)
	no_of_sheets = len(workbook.sheetnames)
	for i in range(0,no_of_sheets):
		if  workbook.sheetnames[i] == sheet_name1:
			worksheet1 = workbook.worksheets[i]
			logging.info(f'Sheet {sheet_name1} of Excel {filename} loaded successfully')
			print((f'Sheet {sheet_name1} of Excel {filename} loaded successfully'))
		elif workbook.sheetnames[i] == sheet_name2:
			worksheet2 = workbook.worksheets[i]
			logging.info(f'Sheet {sheet_name2} of Excel {filename} loaded successfully')
			print((f'Sheet {sheet_name2} of Excel {filename} loaded successfully'))
	return workbook,worksheet1,worksheet2


def read_excel(Wb1_Worksheet1,i):
		B_number = Wb1_Worksheet1.cell(row=i,column=1).value
		charge_description = Wb1_Worksheet1.cell(row=i,column=4).value
		duration = Wb1_Worksheet1.cell(row=i,column=2).value
		dt = str(Wb1_Worksheet1.cell(row=i,column=3).value)
		MSISDN = Wb1_Worksheet1.cell(row =i,column = 7).value
		return B_number,charge_description,duration,dt,MSISDN


def update_results_excel(Wb1_Worksheet1,r,c,value):
	Wb1_Worksheet1.cell(row=r,column=c).value = value


def Validate_date(dt,flag):
	dt = dt[0:10]
	d = str(datetime.date.today())
	if flag.upper() == 'N':
		return True
	else:
		return (dt == d)
	

def check_prefix(B_number):
	if B_number.startswith('00'):
		logging.DEBUG(f'Value of B_number after prefix check is {B_number}')
		return B_number
	else:
		B_number = '00'+B_number
		logging.info(f'Value of B_number after prefix check is {B_number}')
		return B_number

def find_Traffic_Type(charge_description):
	call_period = ''
	Traffic_Type = ''
	special_case = False
	if charge_description == "Voice International":
		Traffic_Type = "Domestic Voice Protocol"
		call_period = 'Per minute'
		return Traffic_Type,call_period,special_case
	elif charge_description == "SMS International":
		Traffic_Type = "Domestic SMS Protocol"
		call_period = 'Per unit'
		return Traffic_Type,call_period,special_case
	elif charge_description == "SMS National Premium":
		Traffic_Type = "Domestic SMS Protocol"
		call_period = 'Per unit'
		special_case = True
		return Traffic_Type,call_period,special_case
	elif charge_description == "SMS National Mobile":
		Traffic_Type = "Domestic SMS Protocol"
		call_period = 'Per unit'
		return Traffic_Type,call_period,special_case
	elif charge_description == "SMS Roaming":
		Traffic_Type = "Roaming Voice Protocol"
		call_period = 'Per minute'
		return Traffic_Type,call_period,special_case
	elif charge_description == "Voice Roaming":
		Traffic_Type = "Roaming SMS Protocol"
		call_period = 'Per unit'
		return Traffic_Type,call_period,special_case
	elif charge_description == "Voice National Premium":
		Traffic_Type = "Domestic Voice Protocol"
		call_period = 'Per minute'
		special_case = True
		return Traffic_Type,call_period,special_case
	elif charge_description == 'Voice National':
		Traffic_Type = "Domestic Voice Protocol"
		call_period = 'Per minute'
		special_case = True
		return Traffic_Type,call_period,special_case
	elif charge_description == 'Voice National Mobile':
		Traffic_Type = "Domestic Voice Protocol"
		call_period = 'Per minute'
		special_case = True
		return Traffic_Type,call_period,special_case
	elif charge_description == 'Voice National Wireline':
		Traffic_Type = "Domestic Voice Protocol"
		call_period = 'Per minute'
		special_case = True
		return Traffic_Type,call_period,special_case
	else :
		raise Exception("No Match for Traffic_Type")


def find_exact_prefix(B_number,worksheet1,rows_worksheet1):
	j=9
	while(j>=3):
		prefix = B_number[0:j]
		for i in range(2,rows_worksheet1+1):
			p = worksheet1.cell(row=i,column=1).value
			if prefix == p:
				return prefix
		j = j-1
	logging.error(f"No match of prefix for B_number: {B_number}")
	return 'NA'


def find_rate_country(prefix,worksheet1,rows_worksheet1,Traffic_Type):
	rate_band='NA'
	Country='NA'
	for i in range(2,rows_worksheet1+1):
		if prefix == worksheet1.cell(row=i,column=1).value and worksheet1.cell(row=i,column=3).value == Traffic_Type:
			rate_band = worksheet1.cell(row=i,column=4).value
			Country = worksheet1.cell(row=i,column=5).value
			return rate_band,Country
	logging.error(f"No Match found for rate and country of prefix : {prefix}")
	return rate_band,Country


def find_charge_Granularity(rate_band,worksheet2,rows_worksheet2,call_period):
	charge =''
	granularity =''
	connection_fees = ''
	for k in range(2,rows_worksheet2+1):
		if rate_band == worksheet2.cell(row=k,column=2).value and call_period == worksheet2.cell(row = k,column=9).value:
			charge = str(worksheet2.cell(row=k,column=8).value)
			granularity = worksheet2.cell(row=k,column=10).value
			connection_fees = worksheet2.cell(row=k,column=6).value
			connection_period = worksheet2.cell(row=k,column=7).value
			connection_period = connection_period[0]
			logging.info(type(charge))
			type(granularity)
			return charge,granularity,connection_fees,connection_period


def Calculate_price_domestic_voice_protocol(duration,ch,gr,connection_fees,connection_period):
	total_charge = 0
	ch=float(ch)
	if connection_fees == '-':
		round_off_duration = math.ceil(duration/gr)
		charge_granularity = ch/60*gr
		total_charge = round_off_duration*charge_granularity
		return total_charge
	else:
		if connection_period == '-':
			connection_period = '0'
		connection_period = int(connection_period)
		connection_period = connection_period*60
		#print(connection_period)
		total_charge += connection_fees
		if duration > connection_period:
			duration -= connection_period
		else:
			duration = 0
		round_off_duration = math.ceil(duration/gr)
		charge_granularity = ch/60*gr
		total_charge += round_off_duration*charge_granularity
		return total_charge

def Calculate_price_domestic_sms_protocol(ch,gr):
		total_charge = ch*gr
		return total_charge

def check_granularity(granularity):
	if(granularity == "1second"):
		g=granularity[0:1]
	elif(granularity == "1minute" or granularity == "1 minute"):
		g='60'
	elif(granularity == "30second"):
		g=granularity[0:2]
	granularity=g
	return granularity



def Domestic_SMS_Protocol(B_number,worksheet1,rows_worksheet1,Traffic_Type,worksheet2,rows_worksheet2,call_period,i,Result):
	prefix = find_exact_prefix(B_number,worksheet1,rows_worksheet1)
	update_results_excel(Result,i,1,prefix)
	logging.info(f'\nExact value of your prefix is {prefix}')
	if prefix == 'NA':
		return
	rate_band,Country = find_rate_country(prefix,worksheet1,rows_worksheet1,Traffic_Type)
	update_results_excel(Result,i,2,rate_band)
	update_results_excel(Result,i,5,Country)
	logging.info(f'\nRate Band of the Called Number is {rate_band}')
	logging.info(f'\nCalled Country For the given number is {Country}\n')
	if (rate_band == 'NA'):
		price = 'NA'
		update_results_excel(Result,i,10,price)
	else:
		charge,granularity,connection_fees,connection_period = find_charge_Granularity(rate_band,worksheet2,rows_worksheet2,call_period)
		update_results_excel(Result,i,8,connection_fees)
		update_results_excel(Result,i,9,connection_period)
		ch = float(charge)
		gr=1
		update_results_excel(Result,i,6,ch)
		update_results_excel(Result,i,7,gr)
		logging.info(f'\nCharge per minute is {ch}')
		logging.info(f'\ngranularity is {gr}')
		price = Calculate_price_domestic_sms_protocol(ch,gr)
		price = round(price,4)
		update_results_excel(Result,i,10,price)
		logging.info('-----------------------------------------------------------------------')
		logging.info(f'\nTotal Charge for the call made is:  {price}')
		logging.info('-----------------------------------------------------------------------')


def Domestic_Voice_Protocol(B_number,worksheet1,rows_worksheet1,Traffic_Type,worksheet2,rows_worksheet2,call_period,i,duration,Result):
	prefix = find_exact_prefix(B_number,worksheet1,rows_worksheet1)
	logging.info(f'\nExact value of your prefix is {prefix}')
	update_results_excel(Result,i,1,prefix)
	if prefix == 'NA':
		return
	rate_band,Country = find_rate_country(prefix,worksheet1,rows_worksheet1,Traffic_Type)
	update_results_excel(Result,i,2,rate_band)
	update_results_excel(Result,i,5,Country)
	logging.info(f'\nRate Band of the Called Number is {rate_band}')
	logging.info(f'\nCalled Country For the given number is {Country}\n')
	if (rate_band == 'NA'):
		price = 'NA'
		update_results_excel(Result,i,10,price)
	else:
		charge,granularity,connection_fees,connection_period = find_charge_Granularity(rate_band,worksheet2,rows_worksheet2,call_period)
		logging.info(f'connection_fees = {connection_fees} and connection_period = {connection_period}')
		update_results_excel(Result,i,8,connection_fees)
		update_results_excel(Result,i,9,connection_period)
		ch = charge[0:8]
		ch = float(ch)
		granularity = check_granularity(granularity)
		gr=int(granularity)
		update_results_excel(Result,i,6,ch)
		update_results_excel(Result,i,7,gr)
		logging.info(f'\nCharge per minute is {ch}')
		logging.info(f'\ngranularity is {gr}')
		price = Calculate_price_domestic_voice_protocol(duration,ch,gr,connection_fees,connection_period)
		price = round(price,4)
		update_results_excel(Result,i,10,price)
		logging.info('-----------------------------------------------------------------------')
		logging.info(f'\nTotal Charge for the call made is:  {price}')
		logging.info('\n-----------------------------------------------------------------------')

def Premium_SMS(prefix,worksheet1,rows_worksheet1,Traffic_Type,worksheet2,rows_worksheet2,call_period,i,Result):
	update_results_excel(Result,i,1,prefix)
	logging.info(f'\nExact value of your prefix is {prefix}')
	rate_band,Country = find_rate_country(prefix,worksheet1,rows_worksheet1,Traffic_Type)
	update_results_excel(Result,i,2,rate_band)
	update_results_excel(Result,i,5,Country)
	logging.info(f'\nRate Band of the Called Number is {rate_band}')
	logging.info(f'\nCalled Country For the given number is {Country}\n')
	if (rate_band == 'NA'):
		price = 'NA'
		update_results_excel(Result,i,10,price)
	else:
		charge,granularity,connection_fees,connection_period = find_charge_Granularity(rate_band,worksheet2,rows_worksheet2,call_period)
		update_results_excel(Result,i,8,connection_fees)
		update_results_excel(Result,i,9,connection_period)
		ch = float(charge)
		gr=1
		update_results_excel(Result,i,6,ch)
		update_results_excel(Result,i,7,gr)
		logging.info(f'\nCharge per minute is {ch}')
		logging.info(f'\ngranularity is {gr}')
		price = Calculate_price_domestic_sms_protocol(ch,gr)
		price = round(price,4)
		update_results_excel(Result,i,10,price)
		logging.info('-----------------------------------------------------------------------')
		logging.info(f'\nTotal Charge for the call made is:  {price}')
		logging.info('-----------------------------------------------------------------------')

def Premium_Voice(prefix,worksheet1,rows_worksheet1,Traffic_Type,worksheet2,rows_worksheet2,call_period,i,duration,Result):
	update_results_excel(Result,i,1,prefix)
	logging.info(f'\nExact value of your prefix is {prefix}')
	rate_band,Country = find_rate_country(prefix,worksheet1,rows_worksheet1,Traffic_Type)
	update_results_excel(Result,i,2,rate_band)
	update_results_excel(Result,i,5,Country)
	logging.info(f'\nRate Band of the Called Number is {rate_band}')
	logging.info(f'\nCalled Country For the given number is {Country}\n')
	if (rate_band == 'NA'):
		price = 'NA'
		update_results_excel(Result,i,10,price)
	else:
		ch,granularity,connection_fees,connection_period = find_charge_Granularity(rate_band,worksheet2,rows_worksheet2,call_period)
		update_results_excel(Result,i,8,connection_fees)
		update_results_excel(Result,i,9,connection_period)
		#ch = float(charge)
		granularity = check_granularity(granularity)
		gr = int(granularity)
		update_results_excel(Result,i,6,ch)
		update_results_excel(Result,i,7,gr)
		logging.info(f'\nCharge per minute is {ch}')
		logging.info(f'\ngranularity is {gr}')
		price = Calculate_price_domestic_voice_protocol(duration,ch,gr,connection_fees,connection_period)
		price = round(price,4)
		update_results_excel(Result,i,10,price)
		logging.info('-----------------------------------------------------------------------')
		logging.info(f'\nTotal Charge for the call made is:  {price}')
		logging.info('-----------------------------------------------------------------------')


def Voice_National_Calls(B_number,Result,i,Traffic_Type,call_period):
	prefix= B_number
	price = 0
	if prefix[0:5] == '35348' or prefix[0:2] == '44':
		Country ='Northern Ireland '
	else:
		Country = 'Ireland'
	update_results_excel(Result,i,1,prefix)
	update_results_excel(Result,i,3,Traffic_Type)
	update_results_excel(Result,i,4,call_period)
	update_results_excel(Result,i,2,'-')
	update_results_excel(Result,i,5,Country)
	update_results_excel(Result,i,10,price)
	update_results_excel(Result,i,8,'-')
	update_results_excel(Result,i,9,'-')

def Select_Rate_Card():
	print(Fore.RED+"1. Red Business\n")
	print(Fore.RED+"2. TEB02\n")
	print(Fore.RED+"3. TEB04\n")
	print(Fore.RED+"4. BW725\n")
	print(Fore.RED+"5. BW284\n")
	print(Fore.RED+"6. SMS_DH01\n")
	rce = int(input("Please Select the number for the rate card against which you want to Validate your records\n"))
	if rce == 1:
		Rate_Card_Excel = 'C:\\Users\\ASHISBHA\\Desktop\\Tariff Validation Tool\\Refrence Rate Cards\\Rate_Card_REDBC.xlsx'
		domestic_allowance = 1
	elif rce == 2:
		Rate_Card_Excel = 'C:\\Users\\ASHISBHA\\Desktop\\Tariff Validation Tool\\Refrence Rate Cards\\Rate_Card_TEB02.xlsx'
		domestic_allowance = 0
	elif rce == 3:
		Rate_Card_Excel = 'C:\\Users\\ASHISBHA\\Desktop\\Tariff Validation Tool\\Refrence Rate Cards\\Rate_Card_TEB04.xlsx'
		domestic_allowance = 0
	elif rce == 4:
		Rate_Card_Excel = 'C:\\Users\\ASHISBHA\\Desktop\\Tariff Validation Tool\\Refrence Rate Cards\\Rate_Card_BW725.xlsx'
		domestic_allowance = 0
	elif rce == 5:
		Rate_Card_Excel = 'C:\\Users\\ASHISBHA\\Desktop\\Tariff Validation Tool\\Refrence Rate Cards\\Rate_Card_BW284.xlsx'
		domestic_allowance = 0
	elif rce == 6:
		Rate_Card_Excel = 'C:\\Users\\ASHISBHA\\Desktop\\Tariff Validation Tool\\Refrence Rate Cards\\Rate_Card_SMS_DDH01.xlsx'
		domestic_allowance = 0
	else:
		print("Please enter the correct choice")

	return Rate_Card_Excel,domestic_allowance

try:
	Rate_Card_File_Name,domestic_allowance = Select_Rate_Card()
	print(Fore.BLUE + "Loading required Excel sheets\n")
	Rate_Card,Tariff_1,PRSMS_rates = load_Excel(Rate_Card_File_Name,"Tariff 1","PRSMS Rates")
	Consumer_Digits,SMS_Voice_Number_list = load_Excel_1('C:/Users/priyeshb/OneDrive - AMDOCS/Backup Folders/Desktop/Python/Python 3/Tariff Validation Tool/Refrence Rate Cards/Consumer_digits.xlsx',"SMS_Voice_Number List (B party)")
	wb1,Input,Result = load_Excel('C:/Users/priyeshb/OneDrive - AMDOCS/Backup Folders/Desktop/Python/Python 3/Tariff Validation Tool/Result/Test_Cases.xlsx',"Input","Result")

	rows_SMS_Voice_Number_list = SMS_Voice_Number_list.max_row
	rows_Tariff_1 = Tariff_1.max_row
	rows_PRSMS_rates = PRSMS_rates.max_row
	rows_Input = Input.max_row
	flag = 'N'
	for i in trange(3,rows_Input+1, desc = 'Processing...',bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.GREEN, Fore.RESET),disable = False):
		#time.sleep(1)
		B_number,charge_description,duration,dt,MSISDN = read_excel(Input,i)
		if Validate_date(dt,flag):
			B_number = str(B_number)
			Traffic_Type,call_period,special_case = find_Traffic_Type(charge_description)
			if special_case:
				if charge_description == "Voice National Premium" :
					prefix = find_exact_prefix(B_number,SMS_Voice_Number_list,rows_SMS_Voice_Number_list)
					if (prefix == 'NA'):
						continue
					update_results_excel(Result,i,3,Traffic_Type)
					update_results_excel(Result,i,4,call_period)
					update_results_excel(Result,i,13,MSISDN)
					Premium_Voice(prefix,SMS_Voice_Number_list,rows_SMS_Voice_Number_list,Traffic_Type,Tariff_1,rows_Tariff_1,call_period,i,duration,Result)
				elif charge_description == "SMS National Premium" :	
					prefix = B_number
					update_results_excel(Result,i,3,Traffic_Type)
					update_results_excel(Result,i,4,call_period)
					update_results_excel(Result,i,13,MSISDN)
					Premium_SMS(prefix,SMS_Voice_Number_list,rows_SMS_Voice_Number_list,Traffic_Type,PRSMS_rates,rows_PRSMS_rates,call_period,i,Result)
				elif charge_description == 'Voice National':
					prefix = B_number
					update_results_excel(Result,i,3,Traffic_Type)
					update_results_excel(Result,i,4,call_period)
					update_results_excel(Result,i,13,MSISDN)
					Premium_Voice(prefix,SMS_Voice_Number_list,rows_SMS_Voice_Number_list,Traffic_Type,Tariff_1,rows_Tariff_1,call_period,i,duration,Result)
				elif charge_description == 'Voice National Mobile':
					if domestic_allowance == 0:
						B_number = check_prefix(B_number)
						logging.info(f'\nUpdated value of B Number is {B_number}')
						update_results_excel(Result,i,3,Traffic_Type)
						update_results_excel(Result,i,4,call_period)
						update_results_excel(Result,i,13,MSISDN)
						logging.info(f'\nYou have Selected traffic type as  {Traffic_Type}')
						logging.info(f'\nThus accordingly your call period is {call_period}')
						Domestic_Voice_Protocol(B_number,SMS_Voice_Number_list,rows_SMS_Voice_Number_list,Traffic_Type,Tariff_1,rows_Tariff_1,call_period,i,duration,Result)
					else:
						Voice_National_Calls(B_number,Result,i,Traffic_Type,call_period)
						update_results_excel(Result,i,13,MSISDN)
				elif charge_description == 'Voice National Wireline':
					if domestic_allowance == 0:
						B_number = check_prefix(B_number)
						logging.info(f'\nUpdated value of B Number is {B_number}')
						update_results_excel(Result,i,3,Traffic_Type)
						update_results_excel(Result,i,4,call_period)
						update_results_excel(Result,i,13,MSISDN)
						logging.info(f'\nYou have Selected traffic type as  {Traffic_Type}')
						logging.info(f'\nThus accordingly your call period is {call_period}')
						Domestic_Voice_Protocol(B_number,SMS_Voice_Number_list,rows_SMS_Voice_Number_list,Traffic_Type,Tariff_1,rows_Tariff_1,call_period,i,duration,Result)
					else:
						Voice_National_Calls(B_number,Result,i,Traffic_Type,call_period)
						update_results_excel(Result,i,13,MSISDN)

			else:
				B_number = check_prefix(B_number)
				logging.info(f'\nUpdated value of B Number is {B_number}')
				update_results_excel(Result,i,3,Traffic_Type)
				update_results_excel(Result,i,4,call_period)
				update_results_excel(Result,i,13,MSISDN)
				logging.info(f'\nYou have Selected traffic type as  {Traffic_Type}')
				logging.info(f'\nThus accordingly your call period is {call_period}')
				if (Traffic_Type == 'Domestic Voice Protocol'):
					Domestic_Voice_Protocol(B_number,SMS_Voice_Number_list,rows_SMS_Voice_Number_list,Traffic_Type,Tariff_1,rows_Tariff_1,call_period,i,duration,Result)
				elif(Traffic_Type == 'Domestic SMS Protocol'):
					Domestic_SMS_Protocol(B_number,SMS_Voice_Number_list,rows_SMS_Voice_Number_list,Traffic_Type,Tariff_1,rows_Tariff_1,call_period,i,Result)

	
	logging.info("--- %s seconds ---" % (time.time() - start_time))
	execution_time = time.time()-start_time
	print(Fore.BLUE + f'\nExecution completed in {execution_time} seconds\nValidate the logs in tariff_validation_tool_{start_time}.log ')
	wb1.save("C:/Users/priyeshb/OneDrive - AMDOCS/Backup Folders/Desktop/Python/Python 3/Tariff Validation Tool/Result/Test_Cases.xlsx")
except Exception as e:
  logging.error(f"Exception occurred {e}", exc_info=True)
